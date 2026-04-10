import os
import time
import shutil
import requests
import warnings
import mimetypes

from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.edge.options import Options

from openpyxl import load_workbook
from rich.console import Console
from rich.text import Text

# =====================================
# CONFIGURACIÓN / CREDENCIALES
# =====================================
# (Opcional) Lee de variables de entorno; si no existen, usa las de abajo.
SCM_USER: str
SCM_PASSWORD: str
FORCEBEAT_USER: str
FORCEBEAT_PASSWORD: str
FORCEBEAT_COMPANY: str
FBT_URL: str
SCM_URL: str

# Ruta de EdgeDriver local
EDGE_DRIVER_PATH: str

# =====================================
# UTILIDADES GLOBALES
# =====================================

LIST_TICKETS = []
COUNT_IMAGE = 0
RETRY_LIMIT = 0
DEFAULT_PATH_CONFIGURATIONS = "configurations.txt"

warnings.filterwarnings("ignore")
console = Console()
width = 104

class ImportData:
    @staticmethod
    def from_excel(file_path: str):
    
        unique_numbers = set()
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, values_only=True):
            value = str(row[0])
            try:
                value = int(float(value))  # tolera "123.0"
                unique_numbers.add(value)
            except (ValueError, TypeError):
                continue
        return list(unique_numbers)


class ProcessText:
    @staticmethod
    def print_color_box(text: str, color: str = "green_yellow"):
        mensaje = f"┌{'─'*width}┐\n│ {text:<{width-1}}│\n└{'─'*width}┘"
        salida = Text(mensaje, style=color)
        console.print(salida)

# =====================================
# CLASE PRINCIPAL FbtService
# =====================================

class FbtService:
    @staticmethod
    def get_service():
      
        return EdgeService(executable_path=EDGE_DRIVER_PATH)

    @staticmethod
    def switch_to_main_frame(driver, wait=None, timeout=12):
      
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
        if wait is None:
            wait = WebDriverWait(driver, timeout)
        iframe = wait.until(EC.presence_of_element_located((By.NAME, "mainFrame")))
        driver.switch_to.frame(iframe)
        return wait

    @staticmethod
    def click_toolbar_close(driver, timeout=10):
      
        wait = WebDriverWait(driver, timeout)
        locators = [
            # Por texto visible "Cerrar"
            (By.XPATH, "//div[contains(@class,'tbi') and .//div[normalize-space()='Cerrar']]"),
            # Por ícono 'chiudi' (chiudi64.png)
            (By.XPATH, "//div[contains(@class,'tbi') and .//img[contains(@src,'chiudi')]]"),
            # Por onclick que llame 'Close'
            (By.XPATH, "//div[@onclick and contains(@onclick,'Close') and contains(@class,'tbi')]"),
            # Por <u>C</u>errrar
            (By.XPATH, "//div[contains(@class,'tbi')][.//u[normalize-space()='C'] and contains(.,'errar')]"),
        ]

        for by, xp in locators:
            try:
                btn = wait.until(EC.element_to_be_clickable((by, xp)))
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.6)
                return True
            except Exception:
                continue
        return False

    @staticmethod
    def return_to_search(driver):
      
        # 1) Cerrar Multimedia/Contenidos
        wait = FbtService.switch_to_main_frame(driver)
        closed_1 = FbtService.click_toolbar_close(driver, timeout=12)
        if not closed_1:
            # Fallback: ESC
            try:
                ActionChains(driver).send_keys(Keys.ESCAPE).perform()
                time.sleep(0.5)
            except Exception:
                pass

        # 2) Reposicionar (el frame suele recargarse tras cerrar)
        wait = FbtService.switch_to_main_frame(driver)

        # 3) Cerrar Detalle
        FbtService.click_toolbar_close(driver, timeout=12)

        # 4) Verificar pantalla de búsqueda
        wait = FbtService.switch_to_main_frame(driver)
        wait.until(EC.presence_of_element_located((By.NAME, "_syXWFMAODLCODICEESTERNO")))
        return True

    @staticmethod
    def login(path_temporal: str, user: str, password: str, company: str):
        try:
            options = EdgeOptions()
            options.use_chromium = True
            options.add_argument("--headless=new")
            options.add_argument("--disable-gpu")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_argument("--no-sandbox")
            options.add_argument("--log-level=3")

            prefs = {
                "download.default_directory": path_temporal,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
            }
            options.add_experimental_option("prefs", prefs)

            #service = FbtService.get_service()
            #driver = webdriver.Edge(service=service, options=options)
            # Configurar opciones de Edge
            options = Options()
            options.add_argument("--headless")          # Ejecutar sin ventana
            options.add_argument("--disable-gpu")       # Recomendado para compatibilidad
            options.add_argument("--no-sandbox")        # Evita errores en algunos entornos
            options.add_argument("--disable-dev-shm-usage")  # Previene problemas de memoria
            options.add_argument("--window-size=1920,1080")  # Tamaño "virtual" del navegador

            # Crear navegador (sin especificar driver)
            driver = webdriver.Edge(options=options)
            driver.implicitly_wait(2)

            driver.get(FBT_URL)
            time.sleep(2)

            # Ingresar a frame y login
            iframe = driver.find_element(By.NAME, "mainFrame")
            driver.switch_to.frame(iframe)
            
            
            driver.find_element(By.ID, "USER").send_keys(user)
            driver.find_element(By.ID, "INPUTPASS").send_keys(password)
            driver.find_element(By.ID, "COMPANY").send_keys(company)
            time.sleep(0.5)

            driver.find_element(By.CLASS_NAME, "enterbutton").click()

            # Asegurar que entró (volver al mainFrame y esperar algo de la app)
            FbtService.switch_to_main_frame(driver)
            return driver

        except Exception as e:
            ProcessText.print_color_box("[X] ERROR AL INGRESAR A FORCEBEAT.", color="red")
            print("Detalle:", e)
            return None

    @staticmethod
    def download_photos(driver, path_temporal: str, path_photos: str, list_photos: list[int]):
      
        try:
            wait = WebDriverWait(driver, 12)

            # Asegurar frame
            FbtService.switch_to_main_frame(driver, wait=wait)

            # Menú: Lista TdC
            wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'Lista TdC')]"))).click()
            time.sleep(0.4)

            # Submenú: Búsqueda TdC
            wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'Búsqueda TdC')]"))).click()
            try:
                # Seleccionar criterio (si aplica a tu pantalla)
                select_element = wait.until(
                    EC.presence_of_element_located((By.NAME, "_lyXWFMAODLID_XWFMTSES"))
                )
                Select(select_element).select_by_value("1000063")
            except Exception as e:
                # Dar click en 'Busqueda TdC' pestaña.
                wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "tab tabOn"))).click()

                # Seleccionar criterio (si aplica a tu pantalla)
                select_element = wait.until(
                    EC.presence_of_element_located((By.NAME, "_lyXWFMAODLID_XWFMTSES"))
                )
                Select(select_element).select_by_value("1000063")
            for i, ticket in enumerate(list_photos, start=1):
                ProcessText.print_color_box(f"[*] DESCARGANDO FOTOS DEL TICKET {ticket}...", color="yellow")

                # Aseguramos estar en el frame correcto antes de tocar el buscador
                FbtService.switch_to_main_frame(driver)

                # Campo de búsqueda por código externo
                element = driver.find_element(By.NAME, "_syXWFMAODLCODICEESTERNO")
                element.clear()
                time.sleep(0.2)
                element.send_keys(str(ticket))
                time.sleep(0.2)

                # Buscar
                driver.find_element(By.XPATH, "//button[@type='SUBMIT']").click()
                time.sleep(1.2)

                try:
                    # Abrir detalle
                    WebDriverWait(driver, 12).until(
                        EC.element_to_be_clickable((By.XPATH, '//button[@class="but butAct" and @tabindex="0"]'))
                    ).click()
                    time.sleep(0.7)

                    # Abrir pestaña 'Contenidos'
                    WebDriverWait(driver, 12).until(
                        EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'Contenidos ')]"))
                    ).click()
                    time.sleep(0.8)
                except Exception:
                    ProcessText.print_color_box(f"[X] TICKET {ticket} SIN SOPORTES O SIN DETALLE", color="red")
                    # Intentar volver a la pantalla de búsqueda para el siguiente
                    try:
                        FbtService.return_to_search(driver)
                    except Exception:
                        pass
                    continue
                
                count = len(driver.find_elements(By.XPATH, "//*[@id='TV-rvRfmu']/div/div/div/table/tbody/tr[not(contains(@class, 'tvRowEmpty'))]")) - 1
                while True:
                    ele = FbtService.safe_find_clickable(driver,'//img[@class="icon Enabled icon_down"]',12)
                    if ele:
                        ele.click()
                        time.sleep(0.8)
                        count += len(driver.find_elements(By.XPATH, "//*[@id='TV-rvRfmu']/div/div/div/table/tbody/tr[not(contains(@class, 'tvRowEmpty'))]")) - 1
                    else:
                        eleUp = FbtService.safe_find_clickable(driver,'//img[@class="icon Enabled icon_upup"]',12)
                        if eleUp:
                            eleUp.click()
                            time.sleep(0.8)
                        break
                ProcessText.print_color_box(f"[*] TOTAL DE FOTOS A DESCARGAR {count}...", color="yellow")
                FbtService.get_files(driver, path_temporal, path_photos, ticket, 0, count)
                        # Volver al buscador para el siguiente ticket
                global COUNT_IMAGE
                global RETRY_LIMIT 
                COUNT_IMAGE = 0
                RETRY_LIMIT = 0
                if ticket in LIST_TICKETS:
                    LIST_TICKETS.remove(ticket)
                ProcessText.print_color_box(f"[OK] SOPORTES DEL TICKET {ticket} DESCARGADOS.", color="green")
                ProcessText.print_color_box("[↩] Regresando al buscador para el siguiente ticket...", color="yellow")
                try:
                    FbtService.return_to_search(driver)
                    ProcessText.print_color_box("[OK] Listo para buscar el siguiente ticket.", color="green")
                except Exception as e:
                    ProcessText.print_color_box("[X] No se pudo volver al buscador. Reintentando ubicar el frame...", color="red")
                    try:
                        FbtService.switch_to_main_frame(driver)
                    except Exception:
                        pass

                time.sleep(0.5)
            driver.quit()
            return True

        except Exception as e:
            ProcessText.print_color_box("[X] ERROR AL DESCARGAR FOTOS.", color="red")
            print("Detalle:", e)
            if RETRY_LIMIT <= 3:
                RETRY_LIMIT += 1
                return FbtService.retryProcess()
            return False
        
    @staticmethod
    def safe_find_clickable(driver, xpath, timeout=8):
        try:
            return WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.XPATH, xpath))
            )
        except TimeoutException:
            return None

    @staticmethod
    def get_files(driver, path_temporal: str, path_photos: str, ticket: int, index: int, count: int):
        
        global COUNT_IMAGE
        global RETRY_LIMIT
        
        bloques = COUNT_IMAGE // 20
        
        index = index - (bloques * 20)
        
        for _ in range(bloques):
            try:
                ele = FbtService.safe_find_clickable(driver,'//img[@class="icon Enabled icon_down"]',12)
                if ele:
                    ele.click()
                    time.sleep(0.8)
            except Exception as e:
                ProcessText.print_color_box("SOPORTE CON ERROR (cambiar pagina)", color="red")
                raise Exception
        
        # Listar filas (tabla de contenidos)
        filas = driver.find_elements(By.XPATH, "//*[@id='TV-rvRfmu']//table/tbody/tr[not(contains(@class, 'tvRowEmpty'))]")

        # Empieza en 2 para saltar cabecera si la hay (como en tu script original)
        for j in range(2 + index, len(filas) + 1):
            columna1 = driver.find_element(
                By.XPATH,
                f"//*[@id='TV-rvRfmu']/div/div/div/table/tbody/tr[{j}]/td[1]"
            )
            ProcessText.print_color_box(f"[↓] DESCARGANDO SOPORTE #{COUNT_IMAGE + 1}/{count}...", color="cyan")

            try:
                ActionChains(driver).double_click(columna1).perform()
                time.sleep(0.8)

                ventanas = driver.window_handles
                if len(ventanas) > 1:
                    driver.switch_to.window(ventanas[-1])

                    # Descargar el recurso de la ventana nueva
                    url = driver.current_url
                    response = requests.get(url)
                    content_type = response.headers.get("Content-Type", "application/octet-stream")
                    extension = mimetypes.guess_extension(content_type) or ".bin"

                    # Intentar obtener filename del header
                    cd = response.headers.get("Content-Disposition", "")
                    if "filename=" in cd:
                        try:
                            raw_name = cd.split("filename=")[-1].strip('"; ')
                            base_name, ext_in_cd = os.path.splitext(raw_name)
                            if ext_in_cd:
                                extension = ext_in_cd
                        except Exception:
                            pass

                    path_file = os.path.join(path_temporal, f"{ticket}_{COUNT_IMAGE + 1}{extension}")
                    with open(path_file, "wb") as f:
                        f.write(response.content)
                    COUNT_IMAGE += 1
                    index += 1
                    # Cerrar ventana de soporte y volver
                    driver.close()
                    driver.switch_to.window(ventanas[0])
                    # Reingresar al frame principal
                    FbtService.switch_to_main_frame(driver)

            except Exception:
                ProcessText.print_color_box("SOPORTE CON ERROR (doble clic/descarga)", color="red")
                raise Exception

        # Mover soportes del ticket a carpeta final (solo archivos de este ticket)
        ticket_folder = os.path.join(path_photos, str(ticket))
        os.makedirs(ticket_folder, exist_ok=True)
        moved = 0
        for file in list(os.listdir(path_temporal)):
            if file.startswith(f"{ticket}_"):
                try:
                    shutil.move(os.path.join(path_temporal, file), os.path.join(ticket_folder, file))
                    moved += 1
                except Exception as e:
                    ProcessText.print_color_box(f"[!] No se pudo mover {file}: {e}", color="red")

        if moved > 0:
            ele = FbtService.safe_find_clickable(driver,'//img[@class="icon Enabled icon_down"]',12)
            if ele:
                ele.click()
                time.sleep(0.8)
                FbtService.get_files(driver,path_temporal,path_photos,ticket,index,count)

        else:
            ProcessText.print_color_box(f"[!] SIN ARCHIVOS PARA EL TICKET {ticket}.", color="yellow")
        return True
    
    @staticmethod
    def retryProcess():
        # Login y descarga
        driver = FbtService.login(path_temporal, FORCEBEAT_USER, FORCEBEAT_PASSWORD, FORCEBEAT_COMPANY)
        if driver:
            return FbtService.download_photos(driver, path_temporal, path_photos, LIST_TICKETS)
        else:
            ProcessText.print_color_box("[X] No se pudo iniciar sesión en ForceBeat.", color="red")
               
    @staticmethod        
    def read_configurations(ruta_archivo="configurations.txt"):
        configurations = {}
        with open(ruta_archivo, "r", encoding="utf-8") as f:
            for linea in f:
                linea = linea.strip()
                # ignorar líneas vacías o comentarios
                if not linea or linea.startswith("#"):
                    continue
                if "=" in linea:
                    clave, valor = linea.split("=", 1)
                    configurations[clave.strip()] = valor.strip()
        return configurations
        
            
    


# =====================================
# MAIN (ejecución local)
# =====================================

if __name__ == "__main__":
    
    path_configurations = input(f"Ruta de archivo configuración (ENTER para usar {DEFAULT_PATH_CONFIGURATIONS}): ").strip()
    
    if not path_configurations:
        path_configurations = DEFAULT_PATH_CONFIGURATIONS
    
    configurations = FbtService.read_configurations(path_configurations)
    
    SCM_USER = os.getenv("SCM_USER",configurations.get("scm_user"))
    SCM_PASSWORD = os.getenv("SCM_PASSWORD",configurations.get("scm_password"))
    FORCEBEAT_USER = os.getenv("FORCEBEAT_USER",configurations.get("forcebeat_user"))
    FORCEBEAT_PASSWORD = os.getenv("FORCEBEAT_PASSWORD",configurations.get("forcebeat_password"))
    FORCEBEAT_COMPANY = os.getenv("FORCEBEAT_COMPANY",configurations.get("forcebeat_company"))
    FBT_URL = os.getenv("FBT_URL",configurations.get("forcebeat_url"))
    SCM_URL = os.getenv("SCM_URL",configurations.get("scm_password"))
    EDGE_DRIVER_PATH = os.getenv("EDGE_DRIVER_PATH",configurations.get("edge_driver_path"))
    
    # Configurar rutas
    path_temporal = os.getenv("TICKETS_PATH_TEMP",os.path.expanduser(configurations.get("temp_path")))
    path_photos   = os.getenv("TICKETS_PATH_PHOTOS",os.path.expanduser(configurations.get("photos_path")))
    
    excel_path = os.getenv("TICKETS_XLSX",configurations.get("xlsx_path"))
    
    if SCM_USER and SCM_PASSWORD and FORCEBEAT_USER and FORCEBEAT_PASSWORD and FORCEBEAT_COMPANY and FBT_URL and SCM_URL and EDGE_DRIVER_PATH and path_temporal and path_photos and excel_path:

        os.makedirs(path_temporal, exist_ok=True)
        os.makedirs(path_photos, exist_ok=True)

        # Leer tickets desde Excel

        tickets = ImportData.from_excel(excel_path)
        if not tickets:
            ProcessText.print_color_box("[!] No se encontraron tickets válidos en el Excel.", color="yellow")
        else:
            ProcessText.print_color_box(f"[+] {len(tickets)} tickets a procesar.", color="green")

            # Login y descarga
            driver = FbtService.login(path_temporal, FORCEBEAT_USER, FORCEBEAT_PASSWORD, FORCEBEAT_COMPANY)
            if driver:
                tickets = sorted(set(tickets))
                LIST_TICKETS = tickets.copy()

                FbtService.download_photos(driver, path_temporal, path_photos, tickets)
            else:
                ProcessText.print_color_box("[X] No se pudo iniciar sesión en ForceBeat.", color="red")
    else:
        ProcessText.print_color_box("[X] No se encontro el archivo de configuración o las variables de entorno.", color="red")

